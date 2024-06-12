import os
import openpyxl

def create_directory_structure(base_dir, sheet_data):
  """
  Creates a directory structure with two subdirectories,
  each containing an excel file with specified headings in multiple tabs.

  Args:
    base_dir: The base directory path where the structure will be created.
    sheet_data (list): A list of dictionaries, where each dictionary represents
                       a sheet and contains keys 'title' for the sheet name and
                       'headings' for a list of column headings.
  """
  sub_dirs = ["Subdirectory 1", "Subdirectory 2"]
  for sub_dir in sub_dirs:
    path = os.path.join(base_dir, sub_dir)
    os.makedirs(path, exist_ok=True)  # Create directory if it doesn't exist

    # Create empty excel file
    filename = f"{sub_dir}.xlsx"
    filepath = os.path.join(path, filename)
    workbook = openpyxl.Workbook()

    # Add sheets with specified headings
    for i, sheet_info in enumerate(sheet_data):
      sheet = workbook.active if i == 0 else workbook.create_sheet()  # Only create sheet on first iteration
      sheet.title = sheet_info["title"]
      for col, heading in enumerate(sheet_info["headings"], start=1):
        sheet.cell(row=1, column=col).value = heading

    workbook.save(filepath)

  print(f"Directory structure created in: {base_dir}")

# Example usage
sheet_data = [
  {"title": "Sheet 1", "headings": ["Column A", "Column B", "Column C"]},
  {"title": "Sheet 2", "headings": ["Data 1", "Data 2"]},
  {"title": "Sheet 3", "headings": ["Measurement 1", "Measurement 2", "Measurement 3"]},
]
base_dir = "your_desired_base_directory"
create_directory_structure(base_dir, sheet_data)
